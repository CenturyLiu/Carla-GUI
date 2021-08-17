#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Jul 22 11:20:28 2020

@author: shijiliu & zhiychen
"""

# define a class for creating the simulation environment for the section

import sys
sys.path.append("..")

import carla
from backend.carla_env import CARLA_ENV 
import math
import time
import numpy as np
from configobj import ConfigObj
import pandas as pd
import copy
from openpyxl import load_workbook
from openpyxl import Workbook

from backend.section_definition import Section
from backend.section_init_definition import InitSection
from backend.generate_path_omit_regulation import generate_path
from backend.intersection_definition import smooth_trajectory, get_trajectory
from backend.multiple_vehicle_control import VehicleControl
from backend.multiple_vehicle_control_debug import VehicleControl_debug
from backend.initial_intersection import get_ego_spectator, get_ego_left_spectator, get_ego_driving_spectator
from backend.section_vehicle_control import VehicleControlFreeway, FullPathVehicleControl, LeadFollowVehicleControl
from backend.human_ego_control import HumanEgoControlServer
import pandas as pd


# color for debug use
red = carla.Color(255, 0, 0)
green = carla.Color(0, 255, 0)
blue = carla.Color(47, 210, 231)
cyan = carla.Color(0, 255, 255)
yellow = carla.Color(255, 255, 0)
orange = carla.Color(255, 162, 0)
white = carla.Color(255, 255, 255)


class FreewayEnv(object):
    # this class holds all methods related to create a simulation environment for freeway
    def __init__(self, env, number_of_sections, max_speed = 30.0, min_speed = 15.0):
        '''
        initialize the freeway simulation environment

        Parameters
        ----------
        number_of_sections : int
            number of sections for the freeway environment.

        max_speed : float
            the max speed of vehicle. Default to be 30.0. 
        
        min_speed : float
            the min speed of vehicle. Default to be 15.0. Vehicle will go under the average of max_speed and min_speed under speed mode
        
        Returns
        -------
        None

        '''
        # debug 
        self.DEBUG_TRAJECTORY = False#True
        self.DEBUG_SUBJECTPOINTS = True
        self.DEBUG_SECTION = True

        # Enable/Disable Data Collection
        self.RECORD_ENABLE = False
        
        # store the number of sections
        self.number_of_sections = number_of_sections
        
        self.env = env
        
        self.spectator = self.env.world.get_spectator()
        
        
        
        self.carla_map = self.env.world.get_map()
        
        # get the subject waypoint of each section
        self.subject_point_list = self._get_section_subject_points((-80.0,9.46))
        
        # store the reference speed
        self.max_speed = max_speed
        self.min_speed = min_speed
        self.navigation_speed = (self.max_speed + self.min_speed) * 0.5
        
        # create the simulation environment and generate the trajectory
        self._create_simulation_env()
        
        # give the trajectory and reference speed list to the initial intersection
        self.section_list[0].get_full_path_trajectory(self.smoothed_full_trajectory, self.ref_speed_list, self.left_smoothed_full_trajectory, self.left_ref_speed_list, self.navigation_speed, self.max_speed, self.min_speed)
    
        
    
    
    #def __del__(self):
    #    self.env.destroy_actors()
    
    # public methods
    def get_section_list(self):
        return self.section_list
    
    def get_vehicle_bounding_box(self,uniquename):
        '''
        get the bounding box of the vehicle by uniquename
        Parameters
        ----------
        uniquename : string
            the uniquename of the vehicle.
        Returns
        -------
        new_bb : carla.Vector3D
            the bounding box of the vehicle, new_bb.x is the length, new_bb.y is the width, new_bb.z is the height
        '''
        new_bb = self.env.get_vehicle_bounding_box(uniquename)
        return new_bb
    
    def add_ego_vehicle(self, model_name = "vehicle.tesla.model3", safety_distance = 15.0, vehicle_color = None):
        # wrapper for add_ego_vehicle function of the init_section
        
        '''
        add ego vehicle to the initial intersection
        according to the user case, the ego vehicle will follow the 
        subject lane all the way with constant speed
        
        Parameters
        ----------
        model_name : string, optional
            vehicle type. The default is "vehicle.tesla.model3".
        safety_distance : float, optional
            smallest distance between this vehicle and vehicle ahead
        vehicle_color : string
            the RGB representation of the vehicle color. e.g. '255,255,255'

        Returns
        -------
        uniquename : string
            the name of the vehicle

        '''
        
        
        uniquename = self.section_list[0].add_ego_vehicle(model_name = model_name, safety_distance = safety_distance, vehicle_color = vehicle_color)
        return uniquename
    
    def edit_ego_vehicle(self, model_name = "vehicle.tesla.model3", safety_distance = 15.0, vehicle_color = None):
        # wrapper for edit_ego_vehicle function of the init_section
        '''
        edit the ego vheicle setting by delete the original ego vehicle and add a new one
        
        Parameters
        ----------
        model_name : string, optional
            vehicle type. The default is "vehicle.tesla.model3".
        safety_distance : float, optional
            smallest distance between this vehicle and vehicle ahead
        vehicle_color : string
            the RGB representation of the vehicle color. e.g. '255,255,255'

        Returns
        -------
        uniquename : string
            the name of the vehicle

        '''
        uniquename = self.section_list[0].edit_ego_vehicle(model_name = model_name, safety_distance = safety_distance, vehicle_color = vehicle_color)
        return uniquename
    
    def add_full_path_vehicle(self, model_name = "vehicle.tesla.model3", vehicle_type ="lead", choice = "subject", command = "speed", command_start_time = 0.0, gap = 10.0, safety_distance = 15.0, lead_follow_distance = 20.0, vehicle_color = None):
        # wrapper for add_full_path_vehicle of the init_section
        '''
        add full path vehicle

        Parameters
        ----------
        model_name : string, optional
            vehicle type. The default is "vehicle.tesla.model3".
        vehicle_type : string, optional
            the vehicle type, valid values : "lead", "follow". The default is "lead".
        choice : string, optional
            the lane choice, valid values are "subject", "left". The default is "subject".
        command : string, optional
            the command the vehicle is going to execute in this section. Valid values: "speed", "lane", "distance". The default is "speed".
        command_start_time : string, optional
            the time at which the command should be executed. The default is 0.0.
        gap : float, optional
            the gap between the vehicle and the one in the front of it when adding. The default is 10.0, unit: meter
        safety_distance : float, optional
            smallest distance between 2 vehicles when simulation is going. The default is 15.0, unit: meter
        vehicle_color : string
            the RGB representation of the vehicle color. e.g. '255,255,255'

        Returns
        -------
        uniquename : string
            the name of the vehicle


        '''
        # add the full path vehicle to the init section
        uniquename = self.section_list[0].add_full_path_vehicle(model_name = model_name, vehicle_type = vehicle_type, choice = choice, command = command, command_start_time = command_start_time, gap = gap, safety_distance = safety_distance, lead_follow_distance = lead_follow_distance, vehicle_color = vehicle_color )
     
        
        # create default vehicle setting for normal sections
        for ii in range(1,len(self.section_list)):
            self.section_list[ii]._add_full_path_vehicle_normal(uniquename, vehicle_type, choice)
    
        return uniquename
    
    
    def remove_full_path_vehicle(self, uniquename):
        '''
        remove a full path vehicle based on its uniquename
        this includes the vehicle setting in init section and 
        settings in other normal sections

        Parameters
        ----------
        uniquename : string
            name of the vehicle.

        Returns
        -------
        removed : Bool
            whether the given vehicle is found and removed

        '''
        removed, vehicle_type, choice, index = self.section_list[0].remove_full_path_vehicle(uniquename)
        if removed:
            for ii in range(1, len(self.section_list)):
                self.section_list[ii]._remove_full_path_vehicle_normal(vehicle_type, choice, index) # remove the setting from normal
                                                                                                    # sections
        
        return removed
        
    def edit_full_path_vehicle_init_setting(self, uniquename, vehicle_type, choice, model_name = "vehicle.tesla.model3",   command = "speed", command_start_time = 0.0, gap = 10.0, safety_distance = 15.0, lead_follow_distance = 20.0, vehicle_color = None):
        # wrapper for the edit_full_path_vehicle function of the init section
        '''
        edit full path vehicle settings by deleting the original vehicle and then add a new one

        Parameters
        ----------
        uniquename : string
            the name of the vehicle
        vehicle_type : string, 
            the vehicle type, valid values : "lead", "follow". 
        choice : string, optional
            the lane choice, valid values are "subject", "left". 
        model_name : string, optional
            vehicle type. The default is "vehicle.tesla.model3".
        command : string, optional
            the command the vehicle is going to execute in this section. Valid values: "speed", "lane", "distance". The default is "speed".
        command_start_time : string, optional
            the time at which the command should be executed. The default is 0.0.
        gap : float, optional
            the gap between the vehicle and the one in the front of it when adding. The default is 10.0, unit: meter
        safety_distance : float, optional
            smallest distance between 2 vehicles when simulation is going. The default is 15.0, unit: meter
        vehicle_color : string
            the RGB representation of the vehicle color. e.g. '255,255,255'

        Returns
        -------
        new_uniquename : string
            the new_name of the vehicle


        '''
        
        new_uniquename, index = self.section_list[0].edit_full_path_vehicle(uniquename = uniquename, 
                                                                     vehicle_type = vehicle_type, 
                                                                     choice = choice,
                                                                     model_name = model_name,
                                                                     command = command,
                                                                     command_start_time = command_start_time,
                                                                     gap = gap, 
                                                                     safety_distance = safety_distance, 
                                                                     lead_follow_distance = lead_follow_distance, 
                                                                     vehicle_color = vehicle_color)
        
        
        
        for ii in range(1,len(self.section_list)):
            self.section_list[ii]._update_vehicle_uniquename(vehicle_type = vehicle_type, choice = choice,index = index,uniquename = new_uniquename)
        
        return new_uniquename
    
    def edit_normal_section_setting(self, section_id, vehicle_type, choice, vehicle_index, command = "speed", command_start_time = 0.0):
        '''
        function for editing vehicle settings for normal section

        Parameters
        ----------
        section_id : TYPE
            the section id. e.g., the first normal section is the second section, so enter section_id = 2
        vehicle_type : TYPE
            DESCRIPTION.
        choice : TYPE
            DESCRIPTION.
        vehicle_index : int
            the index of the vehicle
        command : TYPE, optional
            DESCRIPTION. The default is "speed".
        command_start_time : TYPE, optional
            DESCRIPTION. The default is 0.0.

        Returns
        -------
        None.

        '''
        if section_id > len(self.section_list):
            print("invalid section_id")
            return
        
        section = self.section_list[section_id - 1]
        
        # edit the section settings
        section.edit_full_path_vehicle_local_setting(vehicle_type, choice, vehicle_index, command = command, command_start_time = command_start_time)
    
    def dataOnScreen(self):
        """
        print the ego car data on the screen

        Error: This function is not working as it stuck on run
        it might be helpful to use thread to do async
        """
        world_snapshot = self.env.world.get_snapshot()
        ego_uniquename = self.section_list[0]
        ego_id = (int)(ego_uniquename.split("_")[1])
        ego_actor = world_snapshot.find(ego_id)
        debug = self.env.world.debug

        while True:
            actor_velocity = ego_actor.get_velocity()
            x = round(actor_velocity.x, 2)
            y = round(actor_velocity.y, 2)
            z = round(actor_velocity.z, 2)
            text = "Velocity: x=" + str(x) + " y=" + str(y) + " z=" +str(z) + "(m/s)\n"

            debug.draw_string(location = ego_actor.get_transform().location, 
                            text = text,
                            life_time = 0.2)
            time.sleep(0.2)

    def SectionBackend(self, spectator_mode = None, allow_collision = True, enable_human_control = False):
        '''
        back end function for the freeway

        Parameters
        ----------
        spectator_mode : string, optional
            the spectator mode, valid value is "first_person". The default is None.

        allow_collision : bool, optional
            whether collision is allowed during simulation
            
        enable_human_control : bool, optional
            whether ego vehicle is controlled by human

        Returns
        -------
        None.

        '''
        
        init_section = self.section_list[0]
        ego_vehicle =  VehicleControlFreeway(self.env, init_section.ego_vehicle, self.env.delta_seconds, allow_collision)
        ego_uniquename = init_section.ego_vehicle["uniquename"]
        ego_bb = init_section.ego_vehicle["bounding_box"]
        left_follow_vehicle = []
        subject_follow_vehicle = []
        left_lead_vehicle = []
        subject_lead_vehicle = []
        
        # check whether ego vehicle comes to destination
        ego_end = False
        
        # create vehicle control object
        for vehicle_config in init_section.left_follow_vehicle:
            left_follow_vehicle.append(LeadFollowVehicleControl(self.env, vehicle_config, self.env.delta_seconds,allow_collision))
            
        for vehicle_config in init_section.subject_follow_vehicle:
            subject_follow_vehicle.append(LeadFollowVehicleControl(self.env, vehicle_config, self.env.delta_seconds,allow_collision))
            
        for vehicle_config in init_section.left_lead_vehicle:
            left_lead_vehicle.append(LeadFollowVehicleControl(self.env, vehicle_config, self.env.delta_seconds,allow_collision))
            
        for vehicle_config in init_section.subject_lead_vehicle:
            subject_lead_vehicle.append(LeadFollowVehicleControl(self.env, vehicle_config, self.env.delta_seconds,allow_collision))
            
        
        # store the current section that is functioning
        curr_section = init_section
        self.section_list.pop(0)
        
            # if enable_human_control, get the ego vehicle from the environment
        if enable_human_control:
            human_control_server = HumanEgoControlServer() # create the server for receiving the human command
            spectator_mode = "human_driving"
        
        if self.RECORD_ENABLE:
            # record 1: data on an plain text file
            timestr = time.strftime("%Y%m%d-%H%M%S")
            file = open("../data_collection/FrW" + timestr + ".txt", "w+" )
            
            filename="../data_collection/FrW" + timestr + ".xlsx"
            print("start freeway recordings: ")
            world_snapshot = self.env.world.get_snapshot()
            tm = world_snapshot.timestamp
            file.write("the experiment starts from " + str(round(tm.elapsed_seconds, 3)) + "(seconds)\n")


            # record 2: data on an excel file
            header_row = ['timestamp(sec)']
            for key in self.env.vehicle_dict.keys():
                key = key[8:]
                header_row += [key+ '-lane_id', key+'-location_x(m)', key+'-location_y(m)', key+'location_z(m)', key+'-roll(degrees)', key+'-pitch(degrees)', key+'yaw(degrees)',
                            key+'-velocity_x(m/s)', key+'-velocity_y(m/s)', key+'velocity_z(m/s)', key+'-acceleration_x(m/s2)', key+'-acceleration_y(m/s2)', key+'acceleration_z(m/s2)']
            try:
                wb = load_workbook(filename)
                ws = wb.worksheets[0]
            except FileNotFoundError:
                wb = Workbook()
                ws = wb.active
            ws.append(header_row)
            wb.save(filename)
        
        # main loop for control    
        while True:
            if self.RECORD_ENABLE:
                world_snapshot = self.env.world.get_snapshot()
                ego_id = (int)(ego_uniquename.split("_")[1])
                ego_actor = world_snapshot.find(ego_id)

                map = self.env.world.get_map()
                tm = world_snapshot.timestamp       
                file.write("time: " + str(round(tm.elapsed_seconds, 3))+"(seconds)\n")

                data = [str(round(tm.elapsed_seconds, 3))]
                for key in self.env.vehicle_dict.keys():
                    vehicle_data = []
                    id = (int)(key.split("_")[1])
                    actor = world_snapshot.find(id)
                    if(id == ego_id):
                        file.write("ego id: " + key + "\n")
                    else:
                        file.write("vehicle id: " + key + "\n")
                    
                    actor_location = actor.get_transform().location
                    lane = map.get_waypoint(actor_location).lane_id
                    file.write("lane: " + str(lane) + "\n")
                    vehicle_data+=[lane]

                    actor_transform = actor.get_transform()
                    x = round(actor_transform.location.x, 2)
                    y = round(actor_transform.location.y, 2)
                    z = round(actor_transform.location.z, 2)
                    vehicle_data+=[x, y, z]
                    file.write("location: x=" + str(x) + " y=" + str(y) + " z=" +str(z) + "(meters)\n" )

                    x = round(actor_transform.rotation.roll, 2)
                    y = round(actor_transform.rotation.pitch, 2)
                    z = round(actor_transform.rotation.yaw, 2)
                    vehicle_data+=[x, y, z]
                    file.write("Rotation: roll=" + str(x) + " pitch=" + str(y) + " yaw=" +str(z) + "(degrees)\n")

                    actor_velocity = actor.get_velocity()
                    x = round(actor_velocity.x, 2)
                    y = round(actor_velocity.y, 2)
                    z = round(actor_velocity.z, 2)
                    vehicle_data+=[x, y, z]
                    file.write("Velocity: x=" + str(x) + " y=" + str(y) + " z=" +str(z) + "(m/s)\n")

                    actor_acceleration = actor.get_acceleration()
                    x = round(actor_acceleration.x, 2)
                    y = round(actor_acceleration.y, 2)
                    z = round(actor_acceleration.z, 2)
                    vehicle_data+=[x, y, z]
                    file.write("Acceleration: x=" + str(x) + " y=" + str(y) + " z=" +str(z) + "(m/s2)\n")
                    data+=vehicle_data

                ws.append(data)
                wb.save(filename)

        

            self.env.world.tick()
            # update the distance between vehicles after each tick
            self.env.update_vehicle_distance()
            # change spectator view
            
            if self.env.vehicle_available(ego_uniquename) and spectator_mode == "first_person" :
                 spectator_vehicle_transform = self.env.get_transform_3d(ego_uniquename)
                 spectator_transform = get_ego_spectator(spectator_vehicle_transform,distance = -40)
                 self.spectator.set_transform(spectator_transform)
            
            if self.env.vehicle_available(ego_uniquename) and spectator_mode == "human_driving" :
                 spectator_vehicle_transform = self.env.get_transform_3d(ego_uniquename)
                 spectator_transform = get_ego_driving_spectator(spectator_vehicle_transform, ego_bb)
                 self.spectator.set_transform(spectator_transform)
            
            # section based control
            
            # get ego vehicle transform
            if self.env.vehicle_available(ego_uniquename):
                ego_transform = self.env.get_transform_3d(ego_uniquename)
            
            
            # update the distance with ego vehicle and
            # update the elapsed time of the current section
            local_time = curr_section.tick()
            
            for vehicle in left_follow_vehicle:
                vehicle.update_distance_with_ego(ego_transform) # update the distance with ego vehicle
                vehicle.update_local_time(local_time) # update the local time, change lane or keep distance accordingly
                
            for vehicle in subject_follow_vehicle:
                vehicle.update_distance_with_ego(ego_transform) # update the distance with ego vehicle
                vehicle.update_local_time(local_time) # update the local time, change lane or keep distance accordingly
                
            for vehicle in left_lead_vehicle:
                vehicle.update_distance_with_ego(ego_transform) # update the distance with ego vehicle
                vehicle.update_local_time(local_time) # update the local time, change lane or keep distance accordingly
            
            for vehicle in subject_lead_vehicle:
                vehicle.update_distance_with_ego(ego_transform) # update the distance with ego vehicle
                vehicle.update_local_time(local_time) # update the local time, change lane or keep distance accordingly
            
            
            
            
                
            # apply control to vehicles
            if not ego_end:
                if not enable_human_control:
                    ego_end = ego_vehicle.pure_pursuit_control_wrapper()
                else:
                    # get control from human
                    human_command = human_control_server.get_human_command()
                    # format the command into carla.VehicleControl
                    ego_vehicle_control = carla.VehicleControl(throttle = human_command[0] ,steer=human_command[1],brake = human_command[2])
                    # apply control to the vehicle
                    self.env.apply_vehicle_control(ego_uniquename , ego_vehicle_control)
                    
                    ego_end = ego_vehicle.fake_pure_pursuit_control_wrapper()
            
            '''
            print("--------")
            print("len(left_follow_vehicle) == ", len(left_follow_vehicle))
            print("len(subject_follow_vehicle) == ", len(subject_follow_vehicle))
            print("len(left_lead_vehicle) == ", len(left_lead_vehicle))
            print("len(subject_lead_vehicle) == ", len(subject_lead_vehicle))
            '''
            
            for jj in range(len(left_follow_vehicle) - 1,-1,-1):
                vehicle = left_follow_vehicle[jj]
                end_trajectory = vehicle.pure_pursuit_control_wrapper()
                if end_trajectory:
                    left_follow_vehicle.pop(jj)
                    
            for jj in range(len(subject_follow_vehicle) - 1,-1,-1):
                vehicle = subject_follow_vehicle[jj]
                end_trajectory = vehicle.pure_pursuit_control_wrapper()
                if end_trajectory:
                    subject_follow_vehicle.pop(jj)
                    
            for jj in range(len(left_lead_vehicle) - 1,-1,-1):
                vehicle = left_lead_vehicle[jj]
                end_trajectory = vehicle.pure_pursuit_control_wrapper()
                if end_trajectory:
                    left_lead_vehicle.pop(jj)
                    
            for jj in range(len(subject_lead_vehicle) - 1,-1,-1):
                vehicle = subject_lead_vehicle[jj]
                end_trajectory = vehicle.pure_pursuit_control_wrapper()
                if end_trajectory:
                    subject_lead_vehicle.pop(jj)
            
            if ego_end and len(left_follow_vehicle) == 0 and len(subject_follow_vehicle) == 0 and len(left_lead_vehicle) == 0 and len(subject_lead_vehicle) == 0:
                # exit simulation when all vehicles have arrived at their destination
                break
                
            
            # check whether the subject vehicle is in the next section, if that's the case, change the curr_section
            # and apply section based commands to each vehicle
            if len(self.section_list) > 0: # there still exists unvisited section(s)
                # get the new ego transform
                if self.env.vehicle_available(ego_uniquename):
                    ego_transform = self.env.get_transform_3d(ego_uniquename)
                if self.section_list[0].section_start(ego_transform):
                    curr_section = self.section_list.pop(0) # use the new section; remove this section from wait list
                    
                    # apply section based vehicle settings
                    for ii in range(len(left_follow_vehicle)):
                        vehicle = left_follow_vehicle[ii]
                        command, command_start_time = curr_section.get_full_path_vehicle_local_setting("follow","left",ii)
                        if command != None:
                            vehicle.command = command
                            vehicle.command_start_time = command_start_time
                            
                    for ii in range(len(subject_follow_vehicle)):
                        vehicle = subject_follow_vehicle[ii]
                        command, command_start_time = curr_section.get_full_path_vehicle_local_setting("follow","subject",ii)
                        if command != None:
                            vehicle.command = command
                            vehicle.command_start_time = command_start_time
                            
                    for ii in range(len(left_lead_vehicle)):
                        vehicle = left_lead_vehicle[ii]
                        command, command_start_time = curr_section.get_full_path_vehicle_local_setting("lead","left",ii)
                        if command != None:
                            vehicle.command = command
                            vehicle.command_start_time = command_start_time
                            
                    for ii in range(len(subject_lead_vehicle)):
                        vehicle = subject_lead_vehicle[ii]
                        command, command_start_time = curr_section.get_full_path_vehicle_local_setting("lead","subject",ii)
                        if command != None:
                            vehicle.command = command
                            vehicle.command_start_time = command_start_time
        if self.RECORD_ENABLE:
            file.close()
        
    
    # private methods
    def _create_simulation_env(self):
        # create the freeway environment
        # this function is the counterpart of the "create_intersections" 
        # function for intersection backend
        
        self.section_list = [] # list for all sections 
        self.trajectory_ref_point_list = [] # list for full path trajectory waypoints
        
        initial_section = InitSection(self.env, self.subject_point_list[0]) # create the initial intersection
        self.section_list.append(initial_section) 
        self.trajectory_ref_point_list += initial_section.get_section_trajectory_points()
        
        if self.DEBUG_SECTION:
            trajectory_pt = initial_section.get_section_trajectory_points()
            for pt in trajectory_pt:
                loc1 = pt.transform.location
                self.env.world.debug.draw_point(loc1, size = 0.05, color = red, life_time=0.0, persistent_lines=True)
                
        
        for ii in range(1, self.number_of_sections):
            normal_section = Section(self.env, self.subject_point_list[ii])
            self.section_list.append(normal_section)
            self.trajectory_ref_point_list += normal_section.get_section_trajectory_points()
            if self.DEBUG_SECTION:
                trajectory_pt = normal_section.get_section_trajectory_points()
                for pt in trajectory_pt:
                    loc1 = pt.transform.location
                    self.env.world.debug.draw_point(loc1, size = 0.05, color = red, life_time=0.0, persistent_lines=True)
                    
            
        if self.DEBUG_SUBJECTPOINTS:
            color = green
            for ii in range(len(self.subject_point_list)):
                loc1 = self.subject_point_list[ii].transform.location
                self.env.world.debug.draw_point(loc1, size = 0.1, color = color, life_time=0.0, persistent_lines=True)
        
        
        # connect all reference points for all sections
        full_trajectory = generate_path(self.env, self.trajectory_ref_point_list[0] , self.trajectory_ref_point_list[1], waypoint_separation = 4)
        for ii in range(1, len(self.trajectory_ref_point_list) - 1):
            trajectory = generate_path(self.env, self.trajectory_ref_point_list[ii] , self.trajectory_ref_point_list[ii + 1], waypoint_separation = 4)
            full_trajectory += trajectory[1:]
            
        self.section_subject_trajectory = full_trajectory
        '''
        full_trajectory = []
        for pt in self.trajectory_ref_point_list:
            location = pt.transform.location
            full_trajectory.append((location.x,location.y))
        '''
        
        # create smoothed trajectory and reference speed list
        whole_trajectory = [((pt[0],pt[1]),self.navigation_speed) for pt in full_trajectory]
        
        smoothed_full_trajectory, ref_speed_list = get_trajectory(whole_trajectory)
        self.smoothed_full_trajectory = smoothed_full_trajectory
        self.ref_speed_list = ref_speed_list
        
        if self.DEBUG_TRAJECTORY:
            color = cyan
            for ii in range(1,len(smoothed_full_trajectory)):
                loc1 = carla.Location(x = smoothed_full_trajectory[ii - 1][0], y = smoothed_full_trajectory[ii - 1][1], z = 10.0)
                loc2 = carla.Location(x = smoothed_full_trajectory[ii][0], y = smoothed_full_trajectory[ii][1], z = 10.0)
                self.env.world.debug.draw_arrow(loc1, loc2, thickness = 0.05, arrow_size = 0.1, color = color, life_time=0.0, persistent_lines=True)
    
    
        # get the trajectory for the left lane
        self.left_lane_waypoints_list = []
        for waypoint in self.trajectory_ref_point_list:
            left_waypoint = self._get_left_waypoint(waypoint)
            self.left_lane_waypoints_list.append(left_waypoint)
            
        if self.DEBUG_SECTION:
            for ii in range(1,len(self.left_lane_waypoints_list)):
                loc1 = self.left_lane_waypoints_list[ii - 1].transform.location
                loc2 = self.left_lane_waypoints_list[ii].transform.location
                self.env.world.debug.draw_arrow(loc1, loc2, thickness = 0.1, arrow_size = 0.2, color = blue, life_time=0.0, persistent_lines=True)
        
        
        left_full_trajectory = []
        for pt in self.left_lane_waypoints_list:
            location = pt.transform.location
            left_full_trajectory.append((location.x,location.y))
        
        
        left_whole_trajectory = [((pt[0],pt[1]),self.navigation_speed) for pt in left_full_trajectory]
        
        left_smoothed_full_trajectory, left_ref_speed_list = get_trajectory(left_whole_trajectory)
        self.left_smoothed_full_trajectory = left_smoothed_full_trajectory
        self.left_ref_speed_list = left_ref_speed_list
        
        if self.DEBUG_TRAJECTORY:
            color = blue
            for ii in range(1,len(left_smoothed_full_trajectory)):
                loc1 = carla.Location(x = left_smoothed_full_trajectory[ii - 1][0], y = left_smoothed_full_trajectory[ii - 1][1], z = 10.0)
                loc2 = carla.Location(x = left_smoothed_full_trajectory[ii][0], y = left_smoothed_full_trajectory[ii][1], z = 10.0)
                self.env.world.debug.draw_arrow(loc1, loc2, thickness = 0.05, arrow_size = 0.1, color = color, life_time=0.0, persistent_lines=True)
        
            
        
    def _get_unit_left_vector(self,yaw):
        # get the right vector
        right_yaw = (yaw + 90) % 360
        rad_yaw = math.radians(right_yaw)
        right_vector = [math.cos(rad_yaw),math.sin(rad_yaw)]
        right_vector = right_vector / np.linalg.norm(right_vector)
        return right_vector
    
    
    def _get_left_waypoint(self, curr_waypoint):
        # get the point to the left of the current one
        left_shift = 3.0
        
        curr_location = curr_waypoint.transform.location
        ref_yaw = curr_waypoint.transform.rotation.yaw
        left_vector = self._get_unit_left_vector(ref_yaw)
    
        new_location = carla.Location(x = curr_location.x - left_shift * left_vector[0], y = curr_location.y - left_shift * left_vector[1], z = curr_location.z)
        
        left_waypoint = self.carla_map.get_waypoint(new_location)
        return left_waypoint
    
    def _get_section_subject_points(self, initial_point):
        # get the section subject points
        
        world_pose_list = []
        
        initial_point_raw = carla.Location(initial_point[0], initial_point[1], z = 10.0)
        initial_way_point = self.carla_map.get_waypoint(initial_point_raw)
        
        world_pose_list.append(initial_way_point)
        
        curr_waypoint = initial_way_point
        
        for ii in range(self.number_of_sections - 1):
            
            for jj in range(100):
                next_waypoint = self._get_next_waypoint(curr_waypoint, distance = 4)
                curr_waypoint = next_waypoint
            world_pose_list.append(curr_waypoint)
            
        return world_pose_list
    
    def _get_next_waypoint(self,curr_waypoint,distance = 4):
        '''
        

        Parameters
        ----------
        curr_waypoint : carla.Waypoint
            current waypoint.
        distance : float, optional
            "distance" between current waypoint and target waypoint . The default is 10.

        Returns
        -------
        next_waypoint : carla.Waypoint
            next waypoint, "distance" away from curr_waypoint, in the direction of the current way point
        '''
        forward_vector = curr_waypoint.transform.get_forward_vector()

        location = curr_waypoint.transform.location
        raw_spawn_point = carla.Location(x = location.x + distance * forward_vector.x  , y = location.y + distance * forward_vector.y , z = location.z + 0.1)
        
        next_waypoint = self.carla_map.get_waypoint(raw_spawn_point)
        return next_waypoint
    
def main():
    try:
        
        # create the CARLA_ENV helper 
        client = carla.Client("localhost",2000)
        client.set_timeout(10.0)
        world = client.load_world('Town04') # use Town 04 which contains a freeway
        
        # default the weather to be a fine day
        weather = carla.WeatherParameters(
            cloudiness=10.0,
            precipitation=0.0,
            sun_altitude_angle=90.0)
        world.set_weather(weather)
        
        spectator = world.get_spectator()
        spectator.set_transform(carla.Transform(carla.Location(x=-170, y=-151, z=116.5), carla.Rotation(pitch=-33, yaw= 56.9, roll=0.0)))
        
        # create the environment
        env = CARLA_ENV(world)
        time.sleep(2) # sleep for 2 seconds, wait the initialization to finish
        
        # create a 14 section environment (support up to 7)
        freewayenv = FreewayEnv(env,7)
        
        # add ego vehicle
        freewayenv.add_ego_vehicle()
        freewayenv.edit_ego_vehicle(vehicle_color = '255,255,255')
        
        # add 2 lead vehicle and 2 follow vehicle
        '''
        name0 = freewayenv.add_full_path_vehicle(gap = 20.0, vehicle_type = "lead", choice = "subject", command = "lane", command_start_time = 2.0)
        time.sleep(1.0)
        name0 = freewayenv.add_full_path_vehicle(gap = 20.0, vehicle_type = "lead", choice = "subject", command = "lane", command_start_time = 2.0)
        time.sleep(1.0)
        name1 = freewayenv.add_full_path_vehicle(gap = 20.0, vehicle_type = "lead", choice = "left", command = "lane", command_start_time = 2.0)
        time.sleep(1.0)
        name1 = freewayenv.add_full_path_vehicle(gap = 20.0, vehicle_type = "lead", choice = "left", command = "lane", command_start_time = 2.0)
        time.sleep(1.0)
        name2 = freewayenv.add_full_path_vehicle(vehicle_type = "follow", choice = "subject", command = "lane", command_start_time = 3.0)
        time.sleep(1.0)
        name2 = freewayenv.add_full_path_vehicle(vehicle_type = "follow", choice = "subject", command = "lane", command_start_time = 3.0)
        time.sleep(1.0)
        name3 = freewayenv.add_full_path_vehicle(vehicle_type = "follow", choice = "left", command = "lane", command_start_time = 3.0)
        time.sleep(1.0)
        name3 = freewayenv.add_full_path_vehicle(vehicle_type = "follow", choice = "left", command = "lane", command_start_time = 3.0)
        time.sleep(1.0)
        name4 = freewayenv.add_full_path_vehicle(gap = 20.0, vehicle_type = "lead", choice = "subject", command = "lane", command_start_time = 3.0)
        time.sleep(1.0)
        name5 = freewayenv.add_full_path_vehicle(vehicle_type = "follow", choice = "subject", command = "lane", command_start_time = 0.0)
        time.sleep(1.0)
        
        freewayenv.remove_full_path_vehicle(name4)
        freewayenv.remove_full_path_vehicle(name5)
        
        
        
        
        '''
        name0 = freewayenv.add_full_path_vehicle(gap = 20.0, vehicle_type = "lead", choice = "subject")
        name1 = freewayenv.add_full_path_vehicle(gap = 20.0, vehicle_type = "lead", choice = "left")
        name2 = freewayenv.add_full_path_vehicle(vehicle_type = "follow", choice = "subject")
        name3 = freewayenv.add_full_path_vehicle(vehicle_type = "follow", choice = "left")
        name4 = freewayenv.add_full_path_vehicle(gap = 20.0, vehicle_type = "lead", choice = "subject")
        name5 = freewayenv.add_full_path_vehicle(vehicle_type = "follow", choice = "subject")
        
        # get bounding box
        bb = freewayenv.get_vehicle_bounding_box(name1)
        print("bb.x = %f, bb.y = %f, bb.z = %f" % (bb.x,bb.y,bb.z))
        
        
        # adjust the lead and follow vehicle settings in the third section
        freewayenv.edit_normal_section_setting(section_id = 3, vehicle_type = "lead", choice = "subject", vehicle_index = 0,command = "distance")
        freewayenv.edit_normal_section_setting(section_id = 3, vehicle_type = "lead", choice = "left", vehicle_index = 0,command = "distance")
        freewayenv.edit_normal_section_setting(section_id = 3, vehicle_type = "follow", choice = "subject", vehicle_index = 0,command = "distance")
        freewayenv.edit_normal_section_setting(section_id = 3, vehicle_type = "follow", choice = "left", vehicle_index = 0,command = "distance")
        freewayenv.edit_normal_section_setting(section_id = 3, vehicle_type = "lead", choice = "subject", vehicle_index = 1,command = "distance")
        freewayenv.edit_normal_section_setting(section_id = 3, vehicle_type = "follow", choice = "subject", vehicle_index = 1,command = "distance")
        
        # adjust the lead and follow vehicle settings in the fourth section
        freewayenv.edit_normal_section_setting(section_id = 4, vehicle_type = "lead", choice = "subject", vehicle_index = 0,command = "speed", command_start_time = 0.0)
        freewayenv.edit_normal_section_setting(section_id = 4, vehicle_type = "lead", choice = "left", vehicle_index = 0,command = "speed", command_start_time = 0.0)
        freewayenv.edit_normal_section_setting(section_id = 4, vehicle_type = "follow", choice = "subject", vehicle_index = 0,command = "speed", command_start_time = 0.0)
        freewayenv.edit_normal_section_setting(section_id = 4, vehicle_type = "follow", choice = "left", vehicle_index = 0,command = "speed", command_start_time = 0.0)
        freewayenv.edit_normal_section_setting(section_id = 4, vehicle_type = "lead", choice = "subject", vehicle_index = 1,command = "lane")
        freewayenv.edit_normal_section_setting(section_id = 4, vehicle_type = "follow", choice = "subject", vehicle_index = 1,command = "lane")
        
        # adjust the lead and follow vehicle settings in the sixth section
        freewayenv.edit_normal_section_setting(section_id = 6, vehicle_type = "lead", choice = "subject", vehicle_index = 0,command = "lane")
        freewayenv.edit_normal_section_setting(section_id = 6, vehicle_type = "lead", choice = "left", vehicle_index = 0,command = "lane")
        freewayenv.edit_normal_section_setting(section_id = 6, vehicle_type = "follow", choice = "subject", vehicle_index = 0,command = "lane")
        freewayenv.edit_normal_section_setting(section_id = 6, vehicle_type = "follow", choice = "left", vehicle_index = 0,command = "lane")
        freewayenv.edit_normal_section_setting(section_id = 6, vehicle_type = "lead", choice = "subject", vehicle_index = 1,command = "lane")
        freewayenv.edit_normal_section_setting(section_id = 6, vehicle_type = "follow", choice = "subject", vehicle_index = 1,command = "lane")
        # test remove vehicle
        #freewayenv.remove_full_path_vehicle(name4)
        
        
        
        # test editing vehicle settings
        freewayenv.edit_full_path_vehicle_init_setting(name0, gap = 25.0, vehicle_type = "lead", choice = "subject", vehicle_color = '0,0,0')
        freewayenv.edit_full_path_vehicle_init_setting(name1, gap = 25.0, vehicle_type = "lead", choice = "left", vehicle_color = '255,255,255')
        #freewayenv.edit_full_path_vehicle_init_setting(name2, gap = 25.0, vehicle_type = "follow", choice = "subject", vehicle_color = '0,0,0')
        #freewayenv.edit_full_path_vehicle_init_setting(name3, gap = 25.0, vehicle_type = "follow", choice = "left", vehicle_color = '255,255,255')
        
        
        freewayenv.SectionBackend(spectator_mode = "first_person",allow_collision = False)
    finally:
        time.sleep(10)
        env.destroy_actors()
        
    
if __name__ == '__main__':
    main()